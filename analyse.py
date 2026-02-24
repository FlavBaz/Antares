"""
Script de comparaison des simulations Antares - Scénarios EPR
Usage : python compare_antares.py
Placer tous les fichiers Excel dans le même dossier que ce script,
ou modifier DOSSIER_EXCEL ci-dessous.
"""

import pandas as pd
import openpyxl
from pathlib import Path
import sys

# ============================================================
# CONFIGURATION — modifier ces paramètres si besoin
# ============================================================
DOSSIER_EXCEL = "."   # dossier contenant les fichiers .xlsx

# Nom affiché pour chaque fichier (optionnel) — si vide, utilise le nom du fichier
NOMS_SCENARIOS = {
    "14_reaccteurs.xlsx": "14 EPR (référence)",
    "12_reacteurs.xlsx":  "12 EPR",
    "10_reacteurs.xlsx":  "10 EPR",
}
# ============================================================

# Indicateurs à extraire et comment les agréger
INDICATEURS = [
    # (colonne Antares,                    agrégation, label affiché,              unité)
    ("('LOLD', 'Hours', 'EXP')",          "sum",  "LOLD",                          "h/an"),
    ("('LOLP', '%', 'values')",           "mean", "LOLP",                          "%"),
    ("('UNSP. ENRG', 'MWh', 'EXP')",      "sum",  "Energie non servie",            "MWh"),
    ("('SPIL. ENRG', 'MWh', 'EXP')",      "sum",  "Spillage EnR (curtailment)",    "MWh"),
    ("('CO2 EMIS.', 'Tons', 'EXP')",      "sum",  "Emissions CO2",                 "T"),
    ("('OP. COST', 'Euro', 'EXP')",        "sum",  "Coût opérationnel (OPEX)",      "€"),
    ("('MRG. PRICE', 'Euro', 'EXP')",     "mean", "Prix marginal moyen",           "€/MWh"),
    ("('MRG. PRICE', 'Euro', 'EXP')",     "max",  "Prix marginal max",             "€/MWh"),
    ("('BALANCE', 'MWh', 'EXP')",         "sum",  "Balance nette (exports>0)",     "MWh"),
    ("('LOAD', 'MWh', 'EXP')",            "sum",  "Consommation totale",           "MWh"),
    ("('NUCLEAR', 'MWh', 'EXP')",         "sum",  "Production nucléaire",          "MWh"),
    ("('GAS', 'MWh', 'EXP')",             "sum",  "Production gaz",                "MWh"),
    ("('COAL', 'MWh', 'EXP')",            "sum",  "Production charbon",            "MWh"),
    ("('LIGNITE', 'MWh', 'EXP')",         "sum",  "Production lignite",            "MWh"),
    ("('OIL', 'MWh', 'EXP')",             "sum",  "Production fioul",              "MWh"),
    ("('WIND OFFSHORE', 'MWh', 'EXP')",   "sum",  "Eolien offshore",               "MWh"),
    ("('WIND ONSHORE', 'MWh', 'EXP')",    "sum",  "Eolien onshore",                "MWh"),
    ("('SOLAR PV', 'MWh', 'EXP')",        "sum",  "Solaire PV",                    "MWh"),
    ("('SOLAR ROOFT', 'MWh', 'EXP')",     "sum",  "Solaire toiture",               "MWh"),
    ("('H. ROR', 'MWh', 'EXP')",          "sum",  "Hydraulique fil d'eau",         "MWh"),
    ("('H. STOR', 'MWh', 'EXP')",         "sum",  "Hydraulique stockage",          "MWh"),
]

# Indicateurs dérivés calculés après extraction
def calcul_indicateurs_derives(row):
    """Calcule des indicateurs supplémentaires à partir des bruts."""
    derives = {}
    
    # Part du nucléaire dans la production totale
    prod_totale = (row.get("Production nucléaire [MWh]", 0) +
                   row.get("Production gaz [MWh]", 0) +
                   row.get("Production charbon [MWh]", 0) +
                   row.get("Production lignite [MWh]", 0) +
                   row.get("Eolien offshore [MWh]", 0) +
                   row.get("Eolien onshore [MWh]", 0) +
                   row.get("Solaire PV [MWh]", 0) +
                   row.get("Solaire toiture [MWh]", 0) +
                   row.get("Hydraulique fil d'eau [MWh]", 0) +
                   row.get("Hydraulique stockage [MWh]", 0))
    
    if prod_totale > 0:
        derives["Part nucléaire [%]"] = round(
            row.get("Production nucléaire [MWh]", 0) / prod_totale * 100, 1)
        derives["Part fossile [%]"] = round(
            (row.get("Production gaz [MWh]", 0) +
             row.get("Production charbon [MWh]", 0) +
             row.get("Production lignite [MWh]", 0) +
             row.get("Production fioul [MWh]", 0)) / prod_totale * 100, 1)
    
    # Intensité CO2 de la production
    if prod_totale > 0:
        derives["Intensité CO2 [gCO2/kWh]"] = round(
            row.get("Emissions CO2 [T]", 0) * 1000 / prod_totale, 1)
    
    return derives


def lire_excel(filepath):
    """Lit un fichier Excel Antares et extrait les indicateurs clés."""
    print(f"  Lecture : {filepath.name}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = list(ws.iter_rows(min_row=2, values_only=True))
    df = pd.DataFrame(data, columns=headers)
    
    result = {}
    for col, agg, label, unit in INDICATEURS:
        col_label = f"{label} [{unit}]"
        if col in df.columns:
            series = pd.to_numeric(df[col], errors='coerce')
            if agg == "sum":
                result[col_label] = series.sum()
            elif agg == "mean":
                result[col_label] = series.mean()
            elif agg == "max":
                result[col_label] = series.max()
        else:
            result[col_label] = None
    
    return result


def main():
    dossier = Path(DOSSIER_EXCEL)
    fichiers = sorted(dossier.glob("*.xlsx"))
    
    if not fichiers:
        print(f"❌ Aucun fichier .xlsx trouvé dans : {dossier.resolve()}")
        sys.exit(1)
    
    print(f"\n📂 {len(fichiers)} fichier(s) trouvé(s) dans {dossier.resolve()}\n")
    
    resultats = {}
    for f in fichiers:
        nom = NOMS_SCENARIOS.get(f.name, f.stem.replace("_", " "))
        resultats[nom] = lire_excel(f)
    
    # Tableau de comparaison
    df_result = pd.DataFrame(resultats).T
    
    # Ajout des indicateurs dérivés
    derives_list = []
    for scenario, row in df_result.iterrows():
        derives_list.append(calcul_indicateurs_derives(row.to_dict()))
    df_derives = pd.DataFrame(derives_list, index=df_result.index)
    df_result = pd.concat([df_result, df_derives], axis=1)
    
    # Calcul des deltas par rapport au premier scénario (référence)
    ref = df_result.index[0]
    df_delta = df_result.copy()
    for col in df_result.columns:
        try:
            ref_val = df_result.loc[ref, col]
            if ref_val and ref_val != 0:
                df_delta[col] = ((df_result[col] - ref_val) / abs(ref_val) * 100).round(1)
            else:
                df_delta[col] = None
        except:
            pass
    
    # Export Excel avec mise en forme
    output_path = dossier / "comparaison_scenarios_EPR.xlsx"
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Onglet 1 : Valeurs absolues
        df_result.T.to_excel(writer, sheet_name="Valeurs absolues")
        
        # Onglet 2 : Deltas % vs référence
        df_delta.T.to_excel(writer, sheet_name=f"Delta % vs {ref}")
        
        # Onglet 3 : Synthèse indicateurs clés seulement
        cols_cles = [
            "LOLD [h/an]",
            "Energie non servie [MWh]",
            "Spillage EnR (curtailment) [MWh]",
            "Emissions CO2 [T]",
            "Intensité CO2 [gCO2/kWh]",
            "Coût opérationnel (OPEX) [€]",
            "Prix marginal moyen [€/MWh]",
            "Balance nette (exports>0) [MWh]",
            "Production nucléaire [MWh]",
            "Production gaz [MWh]",
            "Part nucléaire [%]",
            "Part fossile [%]",
        ]
        cols_dispo = [c for c in cols_cles if c in df_result.columns]
        df_result[cols_dispo].T.to_excel(writer, sheet_name="Synthèse clés")
        
        # Mise en forme colorée
        wb_out = writer.book
        for sheet_name in wb_out.sheetnames:
            ws_out = wb_out[sheet_name]
            ws_out.column_dimensions['A'].width = 35
            for col_letter in ['B','C','D','E','F','G','H']:
                ws_out.column_dimensions[col_letter].width = 18
    
    print(f"\n✅ Fichier de comparaison généré : {output_path}")
    print(f"\n{'='*60}")
    print("SYNTHÈSE RAPIDE")
    print('='*60)
    
    # Affichage console
    cols_affich = [
        "LOLD [h/an]", "Energie non servie [MWh]",
        "Spillage EnR (curtailment) [MWh]", "Emissions CO2 [T]",
        "Coût opérationnel (OPEX) [€]", "Prix marginal moyen [€/MWh]",
        "Balance nette (exports>0) [MWh]", "Production nucléaire [MWh]",
        "Production gaz [MWh]", "Part nucléaire [%]", "Intensité CO2 [gCO2/kWh]"
    ]
    cols_dispo_aff = [c for c in cols_affich if c in df_result.columns]
    
    with pd.option_context('display.max_columns', None, 'display.width', 200,
                           'display.float_format', '{:,.1f}'.format):
        print(df_result[cols_dispo_aff].T.to_string())


if __name__ == "__main__":
    main()